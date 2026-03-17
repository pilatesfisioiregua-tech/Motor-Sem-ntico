#!/usr/bin/env python3
"""
Agente semántico para archivos locales.
Lee documentos de cualquier directorio, analiza su contenido con LLM,
y propone reorganización basada en semántica.

Uso:
  python3 local_semantic_agent.py --analyze ~/omni-mind-cerebro
  python3 local_semantic_agent.py --analyze ~/Library/Mobile\ Documents/com~apple~CloudDocs
  python3 local_semantic_agent.py --analyze ~/omni-mind-cerebro ~/Library/Mobile\ Documents/com~apple~CloudDocs
  python3 local_semantic_agent.py --propose
  python3 local_semantic_agent.py --execute
  python3 local_semantic_agent.py --map ~/omni-mind-cerebro   # Solo mapa visual semántico
"""

import argparse
import io
import json
import os
import re
import shutil
import subprocess
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / "semantic_data"

OPENROUTER_KEY = "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"

# Modelo barato para análisis masivo
ANALYSIS_MODEL = "deepseek/deepseek-chat-v3-0324"
# Modelo potente para propuesta
PROPOSAL_MODEL = "deepcogito/cogito-v2.1-671b"

MAX_TEXT_CHARS = 3000

# Extensiones que podemos leer
READABLE_EXTENSIONS = {
    # Documentos
    ".pdf", ".docx", ".doc", ".xlsx", ".xls",
    # Texto
    ".txt", ".md", ".csv", ".json", ".yaml", ".yml",
    # Código
    ".py", ".js", ".ts", ".html", ".css", ".sql", ".sh",
    # Config
    ".toml", ".ini", ".cfg", ".env",
}

# Extensiones a ignorar (binarios, media, caches)
SKIP_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif", ".heic", ".heif", ".webp", ".svg",
    ".mp4", ".mov", ".avi", ".mkv", ".m4a", ".mp3", ".wav",
    ".dmg", ".exe", ".msi", ".pkg", ".zip", ".tar", ".gz", ".rar",
    ".pyc", ".pyo", ".class", ".o", ".so", ".dylib",
    ".db", ".sqlite", ".pickle", ".pkl",
    ".DS_Store", ".gitignore",
}

# Directorios a ignorar
SKIP_DIRS = {
    ".git", "__pycache__", "node_modules", ".venv", "venv",
    ".mypy_cache", ".pytest_cache", ".tox", "dist", "build",
    ".egg-info", ".claude", ".next", ".vercel",
    "semantic_data",  # Nuestros propios datos
}


# ─── TEXT EXTRACTION ──────────────────────────────────────────────────────────

def extract_text_pdf(filepath: Path) -> str:
    try:
        import pdfplumber
        text = ""
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages[:10]:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
                if len(text) > MAX_TEXT_CHARS:
                    break
        return text[:MAX_TEXT_CHARS]
    except Exception:
        return ""


def extract_text_docx(filepath: Path) -> str:
    try:
        from docx import Document
        doc = Document(str(filepath))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return text[:MAX_TEXT_CHARS]
    except Exception:
        return ""


def extract_text_xlsx(filepath: Path) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(filepath), read_only=True, data_only=True)
        text = ""
        for sheet in wb.sheetnames[:3]:
            ws = wb[sheet]
            text += f"[Hoja: {sheet}]\n"
            for row in ws.iter_rows(max_row=50, values_only=True):
                vals = [str(v) for v in row if v is not None]
                if vals:
                    text += " | ".join(vals) + "\n"
                if len(text) > MAX_TEXT_CHARS:
                    break
        return text[:MAX_TEXT_CHARS]
    except Exception:
        return ""


def extract_text_plain(filepath: Path) -> str:
    try:
        text = filepath.read_text(encoding="utf-8", errors="replace")
        return text[:MAX_TEXT_CHARS]
    except Exception:
        return ""


EXTRACTORS = {
    ".pdf": extract_text_pdf,
    ".docx": extract_text_docx,
    ".doc": extract_text_docx,
    ".xlsx": extract_text_xlsx,
    ".xls": extract_text_xlsx,
    ".txt": extract_text_plain,
    ".md": extract_text_plain,
    ".csv": extract_text_plain,
    ".json": extract_text_plain,
    ".yaml": extract_text_plain,
    ".yml": extract_text_plain,
    ".py": extract_text_plain,
    ".js": extract_text_plain,
    ".ts": extract_text_plain,
    ".html": extract_text_plain,
    ".css": extract_text_plain,
    ".sql": extract_text_plain,
    ".sh": extract_text_plain,
    ".toml": extract_text_plain,
    ".ini": extract_text_plain,
    ".cfg": extract_text_plain,
    ".env": extract_text_plain,
}


# ─── FILE DISCOVERY ──────────────────────────────────────────────────────────

def discover_files(roots: List[Path]) -> List[dict]:
    """Walk directories and find all readable files."""
    files = []

    for root in roots:
        root = root.resolve()
        if not root.exists():
            print(f"  ! No existe: {root}")
            continue

        for dirpath, dirnames, filenames in os.walk(root):
            # Skip hidden/cache dirs
            dirnames[:] = [d for d in dirnames if d not in SKIP_DIRS and not d.startswith(".")]

            for filename in filenames:
                filepath = Path(dirpath) / filename
                ext = filepath.suffix.lower()

                if ext in SKIP_EXTENSIONS or filename.startswith("."):
                    continue
                if ext not in READABLE_EXTENSIONS:
                    continue

                try:
                    stat = filepath.stat()
                    # Skip files > 50MB
                    if stat.st_size > 50 * 1024 * 1024:
                        continue
                    # Skip empty files
                    if stat.st_size == 0:
                        continue

                    rel_path = str(filepath.relative_to(root))
                    files.append({
                        "path": str(filepath),
                        "rel_path": rel_path,
                        "root": str(root),
                        "name": filename,
                        "ext": ext,
                        "size": stat.st_size,
                        "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                        "dir": str(Path(dirpath).relative_to(root)),
                    })
                except (OSError, ValueError):
                    continue

    return files


# ─── LLM ──────────────────────────────────────────────────────────────────────

def call_llm(model: str, system: str, user_msg: str, max_tokens: int = 500) -> Optional[str]:
    payload = json.dumps({
        "model": model,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user_msg},
        ],
        "max_tokens": max_tokens,
        "temperature": 0.3,
    })

    result = subprocess.run(
        [
            "curl", "-s", "https://openrouter.ai/api/v1/chat/completions",
            "-H", "Content-Type: application/json",
            "-H", f"Authorization: Bearer {OPENROUTER_KEY}",
            "-d", payload,
        ],
        capture_output=True, text=True, timeout=120,
    )

    try:
        data = json.loads(result.stdout)
        content = data["choices"][0]["message"]["content"]
        content = re.sub(r"<think>.*?</think>", "", content, flags=re.DOTALL).strip()
        return content
    except Exception:
        return None


# ─── PHASE 1+2: ANALYZE ──────────────────────────────────────────────────────

ANALYSIS_SYSTEM = """Eres un archivista experto. Analiza el contenido de este archivo y responde SOLO en JSON:

{
  "tema": "tema principal en 3-5 palabras",
  "tags": ["tag1", "tag2", "tag3"],
  "tipo": "código|documentación|config|datos|diseño|negocio|formación|personal|legal|financiero|investigación|otro",
  "relaciones": ["nombre de otro archivo o concepto con el que se relaciona"],
  "resumen": "una frase describiendo el contenido y propósito del archivo"
}"""


def analyze_files(roots: List[Path]) -> None:
    """Phase 1+2: Discover and analyze all files."""
    data_file = DATA_DIR / "analysis_local.json"

    root_names = [r.name for r in roots]
    print(f"\n{'='*60}")
    print(f" Análisis semántico local: {', '.join(root_names)}")
    print(f"{'='*60}")

    # Load existing
    existing = {}
    if data_file.exists():
        with open(data_file) as f:
            existing = json.load(f)
        print(f"  Análisis previo: {len(existing)} archivos")

    # Discover files
    print(f"  Buscando archivos...")
    files = discover_files(roots)
    print(f"  Archivos encontrados: {len(files)}")

    # Stats
    by_ext = defaultdict(int)
    by_root = defaultdict(int)
    for f in files:
        by_ext[f["ext"]] += 1
        by_root[Path(f["root"]).name] += 1

    print(f"\n  Por directorio:")
    for root, count in sorted(by_root.items()):
        print(f"    {root}: {count} archivos")
    print(f"\n  Por tipo: {', '.join(f'{ext}({n})' for ext, n in sorted(by_ext.items(), key=lambda x: -x[1])[:10])}")

    # Analyze
    new_count = 0
    skip_count = 0
    error_count = 0

    for i, file_info in enumerate(files, 1):
        filepath = file_info["path"]

        # Skip if already analyzed (by path)
        if filepath in existing:
            # Check if file was modified since last analysis
            prev_mod = existing[filepath].get("modified", "")
            if prev_mod == file_info["modified"]:
                skip_count += 1
                continue

        ext = file_info["ext"]
        extractor = EXTRACTORS.get(ext)
        if not extractor:
            continue

        text = extractor(Path(filepath))
        if not text.strip():
            error_count += 1
            continue

        prompt = (
            f"Archivo: {file_info['name']}\n"
            f"Ruta: {file_info['rel_path']}\n"
            f"Directorio raíz: {Path(file_info['root']).name}\n\n"
            f"Contenido:\n{text}"
        )

        print(f"  [{i}/{len(files)}] {file_info['rel_path'][:60]}...", end=" ", flush=True)

        llm_response = call_llm(ANALYSIS_MODEL, ANALYSIS_SYSTEM, prompt)

        if llm_response:
            try:
                json_match = re.search(r'\{[^{}]*\}', llm_response, re.DOTALL)
                if json_match:
                    analysis = json.loads(json_match.group())
                else:
                    analysis = {"tema": "parse error", "tags": [], "tipo": "otro", "resumen": llm_response[:100]}
            except json.JSONDecodeError:
                analysis = {"tema": "parse error", "tags": [], "tipo": "otro", "resumen": llm_response[:100]}

            existing[filepath] = {
                **file_info,
                "analysis": analysis,
            }
            new_count += 1
            print(f"→ {analysis.get('tipo', '?')} [{', '.join(analysis.get('tags', [])[:3])}]")
        else:
            error_count += 1
            print("→ ERROR")

        # Save every 10
        if new_count % 10 == 0 and new_count > 0:
            with open(data_file, "w") as f:
                json.dump(existing, f, indent=2, ensure_ascii=False)

        time.sleep(0.3)

    # Final save
    with open(data_file, "w") as f:
        json.dump(existing, f, indent=2, ensure_ascii=False)

    print(f"\n── Resumen ──")
    print(f"  Nuevos: {new_count}  |  Existentes: {skip_count}  |  Errores: {error_count}")

    # Semantic stats
    by_type = defaultdict(int)
    all_tags = defaultdict(int)
    for doc in existing.values():
        a = doc.get("analysis", {})
        by_type[a.get("tipo", "otro")] += 1
        for tag in a.get("tags", []):
            all_tags[tag.lower()] += 1

    print(f"\n  Por tipo semántico:")
    for tipo, count in sorted(by_type.items(), key=lambda x: -x[1]):
        print(f"    {tipo}: {count}")

    print(f"\n  Top tags:")
    for tag, count in sorted(all_tags.items(), key=lambda x: -x[1])[:15]:
        print(f"    #{tag}: {count}")

    print(f"\n  Datos: {data_file}")


# ─── PHASE 3: SEMANTIC MAP ───────────────────────────────────────────────────

MAP_SYSTEM = """Eres un arquitecto de información. Te doy una lista de archivos con su análisis semántico de un repositorio de trabajo.

Genera un MAPA SEMÁNTICO que muestre:

1. **CLUSTERS TEMÁTICOS**: Agrupa los archivos por tema/función (no por carpeta actual). Nombra cada cluster.

2. **CONEXIONES**: Identifica archivos que están relacionados pero en carpetas diferentes.

3. **PROBLEMAS DETECTADOS**:
   - Archivos duplicados o redundantes (mismo tema, diferente versión)
   - Archivos huérfanos (no encajan con ningún cluster)
   - Archivos mal ubicados (su contenido no corresponde a su carpeta)

4. **PROPUESTA DE REORGANIZACIÓN**: Estructura óptima de carpetas con máximo 3 niveles.

Responde en formato markdown legible."""


REORG_SYSTEM = """Eres un arquitecto de información. Te doy archivos con su análisis semántico.

Propone una reorganización CONCRETA en JSON:

{
  "estructura": [
    {
      "carpeta": "nombre/subcarpeta",
      "descripcion": "qué contiene",
      "archivos": [
        {"path_actual": "/ruta/actual/archivo.md", "razon": "por qué va aquí"}
      ]
    }
  ],
  "eliminar": [
    {"path": "/ruta/archivo", "razon": "duplicado de X / obsoleto / vacío"}
  ],
  "renombrar": [
    {"path_actual": "/ruta/actual", "nombre_nuevo": "mejor_nombre.md", "razon": ""}
  ]
}

Reglas:
- No mover archivos de código que están en su lugar correcto (src/, tests/, etc.)
- Consolidar documentos duplicados (quedarse con la versión más reciente)
- Los archivos de configuración (.toml, .env, etc.) se quedan donde están
- Priorizar claridad sobre profundidad (máx 3 niveles)"""


def generate_map(roots: List[Path]) -> None:
    """Generate semantic map of all analyzed files."""
    data_file = DATA_DIR / "analysis_local.json"
    map_file = DATA_DIR / "semantic_map.md"

    if not data_file.exists():
        print("ERROR: Primero ejecuta --analyze")
        return

    with open(data_file) as f:
        analysis = json.load(f)

    # Filter to requested roots
    root_strs = [str(r.resolve()) for r in roots]
    filtered = {}
    for path, doc in analysis.items():
        doc_root = doc.get("root", "")
        if any(doc_root.startswith(r) or r.startswith(doc_root) for r in root_strs) or not root_strs:
            filtered[path] = doc

    if not filtered:
        filtered = analysis  # Use all if no root filter matches

    print(f"\n{'='*60}")
    print(f" Mapa semántico: {len(filtered)} archivos")
    print(f"{'='*60}")

    # Build summaries
    summaries = []
    for path, doc in filtered.items():
        a = doc.get("analysis", {})
        summaries.append(
            f"- {doc.get('rel_path', path)}\n"
            f"  Carpeta: {doc.get('dir', '/')}\n"
            f"  Tipo: {a.get('tipo', '?')} | Tags: {', '.join(a.get('tags', []))}\n"
            f"  Tema: {a.get('tema', '?')}\n"
            f"  Resumen: {a.get('resumen', '?')}"
        )

    all_text = "\n".join(summaries)
    if len(all_text) > 90000:
        all_text = all_text[:90000] + "\n\n[... truncado]"

    print(f"  Enviando a {PROPOSAL_MODEL}...")

    response = call_llm(
        PROPOSAL_MODEL,
        MAP_SYSTEM,
        f"Archivos a analizar:\n\n{all_text}",
        max_tokens=5000,
    )

    if response:
        with open(map_file, "w") as f:
            f.write(f"# Mapa Semántico — {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n")
            f.write(response)

        print(f"\n{response}")
        print(f"\n  Mapa guardado: {map_file}")
    else:
        print("  ERROR: No se obtuvo respuesta")


def propose_reorganization() -> None:
    """Propose concrete reorganization."""
    data_file = DATA_DIR / "analysis_local.json"
    proposal_file = DATA_DIR / "proposal_local.json"

    if not data_file.exists():
        print("ERROR: Primero ejecuta --analyze")
        return

    with open(data_file) as f:
        analysis = json.load(f)

    print(f"\n{'='*60}")
    print(f" Propuesta de reorganización: {len(analysis)} archivos")
    print(f"{'='*60}")

    # Build summaries
    summaries = []
    for path, doc in analysis.items():
        a = doc.get("analysis", {})
        summaries.append(
            f"- PATH: {path}\n"
            f"  Rel: {doc.get('rel_path', '')}\n"
            f"  Dir: {doc.get('dir', '/')}\n"
            f"  Tipo: {a.get('tipo', '?')} | Tags: {', '.join(a.get('tags', []))}\n"
            f"  Resumen: {a.get('resumen', '?')}"
        )

    all_text = "\n".join(summaries)
    if len(all_text) > 90000:
        all_text = all_text[:90000]

    print(f"  Enviando a {PROPOSAL_MODEL}...")

    response = call_llm(
        PROPOSAL_MODEL,
        REORG_SYSTEM,
        f"Archivos:\n\n{all_text}",
        max_tokens=5000,
    )

    if not response:
        print("  ERROR: sin respuesta")
        return

    # Parse JSON
    try:
        json_match = re.search(r'\{[\s\S]*\}', response)
        if json_match:
            proposal = json.loads(json_match.group())
        else:
            proposal = {"raw": response}
    except json.JSONDecodeError:
        proposal = {"raw": response}

    with open(proposal_file, "w") as f:
        json.dump(proposal, f, indent=2, ensure_ascii=False)

    # Display
    if "estructura" in proposal:
        print(f"\n── Estructura propuesta ──\n")
        for folder in proposal["estructura"]:
            carpeta = folder.get("carpeta", "?")
            desc = folder.get("descripcion", "")
            archivos = folder.get("archivos", [])
            print(f"  📁 {carpeta} ({len(archivos)} archivos)")
            if desc:
                print(f"     {desc}")
            for a in archivos[:5]:
                print(f"     ← {Path(a.get('path_actual', '')).name}: {a.get('razon', '')}")
            if len(archivos) > 5:
                print(f"     ... y {len(archivos) - 5} más")

        if proposal.get("eliminar"):
            print(f"\n  🗑️  A eliminar: {len(proposal['eliminar'])}")
            for d in proposal["eliminar"][:5]:
                print(f"     {Path(d.get('path', '')).name}: {d.get('razon', '')}")

        if proposal.get("renombrar"):
            print(f"\n  ✏️  A renombrar: {len(proposal['renombrar'])}")
            for r in proposal["renombrar"][:5]:
                print(f"     {Path(r.get('path_actual', '')).name} → {r.get('nombre_nuevo', '')}")
    else:
        print(response)

    print(f"\n  Propuesta: {proposal_file}")


# ─── PHASE 4: EXECUTE ────────────────────────────────────────────────────────

def execute_reorganization() -> None:
    """Execute proposed reorganization on local filesystem."""
    proposal_file = DATA_DIR / "proposal_local.json"

    if not proposal_file.exists():
        print("ERROR: Primero ejecuta --propose")
        return

    with open(proposal_file) as f:
        proposal = json.load(f)

    if "estructura" not in proposal:
        print("ERROR: Propuesta sin estructura válida")
        return

    total_moves = sum(len(f.get("archivos", [])) for f in proposal["estructura"])
    total_deletes = len(proposal.get("eliminar", []))
    total_renames = len(proposal.get("renombrar", []))

    print(f"\n{'='*60}")
    print(f" Ejecutar reorganización local")
    print(f" Mover: {total_moves} | Eliminar: {total_deletes} | Renombrar: {total_renames}")
    print(f"{'='*60}")

    confirm = input("\n  ⚠️  Esto modificará archivos. ¿Continuar? (s/n): ").strip().lower()
    if confirm != "s":
        print("  Cancelado.")
        return

    moved = 0
    errors = 0

    # Move files
    for folder_spec in proposal["estructura"]:
        carpeta = folder_spec.get("carpeta", "")
        archivos = folder_spec.get("archivos", [])

        for file_spec in archivos:
            src = Path(file_spec.get("path_actual", ""))
            if not src.exists():
                errors += 1
                continue

            # Determine destination
            # Use the first root as base for the new structure
            dest_dir = src.parent.parent / carpeta  # Relative to parent
            dest_dir.mkdir(parents=True, exist_ok=True)
            dest = dest_dir / src.name

            if dest == src:
                continue  # Already in place

            if dest.exists():
                stem, ext = dest.stem, dest.suffix
                counter = 1
                while dest.exists():
                    dest = dest_dir / f"{stem}_{counter}{ext}"
                    counter += 1

            try:
                shutil.move(str(src), str(dest))
                moved += 1
                print(f"  ✓ {src.name} → {carpeta}/")
            except Exception as e:
                errors += 1
                print(f"  ✗ {src.name}: {e}")

    # Renames
    renamed = 0
    for rename_spec in proposal.get("renombrar", []):
        src = Path(rename_spec.get("path_actual", ""))
        new_name = rename_spec.get("nombre_nuevo", "")
        if src.exists() and new_name:
            dest = src.parent / new_name
            if not dest.exists():
                try:
                    src.rename(dest)
                    renamed += 1
                    print(f"  ✏️  {src.name} → {new_name}")
                except Exception as e:
                    print(f"  ✗ Renombrar {src.name}: {e}")

    print(f"\n── Resumen ──")
    print(f"  Movidos:     {moved}")
    print(f"  Renombrados: {renamed}")
    print(f"  Errores:     {errors}")
    print(f"  (Eliminaciones no se ejecutan automáticamente por seguridad)")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Agente semántico para archivos locales")
    parser.add_argument("--analyze", action="store_true", help="Fase 1+2: descubrir y analizar archivos")
    parser.add_argument("--map", action="store_true", help="Generar mapa semántico visual")
    parser.add_argument("--propose", action="store_true", help="Fase 3: proponer reorganización")
    parser.add_argument("--execute", action="store_true", help="Fase 4: ejecutar reorganización")
    parser.add_argument("paths", nargs="*", default=[], help="Directorios a analizar")
    args = parser.parse_args()

    if not any([args.analyze, args.map, args.propose, args.execute]):
        print("Agente semántico para archivos locales")
        print()
        print("Uso:")
        print("  python3 local_semantic_agent.py --analyze ~/omni-mind-cerebro")
        print("  python3 local_semantic_agent.py --analyze ~/omni-mind-cerebro ~/Library/Mobile\\ Documents/com~apple~CloudDocs")
        print("  python3 local_semantic_agent.py --map ~/omni-mind-cerebro")
        print("  python3 local_semantic_agent.py --propose")
        print("  python3 local_semantic_agent.py --execute")
        return

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # Default paths
    roots = [Path(p).expanduser() for p in args.paths] if args.paths else []

    if args.analyze:
        if not roots:
            print("ERROR: Especifica al menos un directorio")
            print("  python3 local_semantic_agent.py --analyze ~/omni-mind-cerebro")
            return
        analyze_files(roots)

    if args.map:
        generate_map(roots)

    if args.propose:
        propose_reorganization()

    if args.execute:
        execute_reorganization()


if __name__ == "__main__":
    main()
