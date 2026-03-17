#!/usr/bin/env python3
"""
Agente semántico de Google Drive.
Lee todos los documentos, analiza su contenido con LLM,
y propone una reorganización basada en semántica (no títulos).

Uso:
  python3 gdrive_semantic_agent.py --analyze                # Fase 1+2: crawl + analizar
  python3 gdrive_semantic_agent.py --propose                # Fase 3: proponer reorganización
  python3 gdrive_semantic_agent.py --execute                # Fase 4: ejecutar (con confirmación)
  python3 gdrive_semantic_agent.py --account gurufe         # Solo una cuenta
"""

import argparse
import base64
import io
import json
import os
import re
import subprocess
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

SCOPES = ["https://www.googleapis.com/auth/drive"]
SCRIPT_DIR = Path(__file__).parent
CREDENTIALS_FILE = SCRIPT_DIR / "credentials.json"
DATA_DIR = SCRIPT_DIR / "semantic_data"

OPENROUTER_KEY = "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"

# Modelo barato para análisis masivo de documentos (~$0.001/doc)
ANALYSIS_MODEL = "deepseek/deepseek-chat-v3-0324"
# Modelo potente para propuesta de reorganización
PROPOSAL_MODEL = "deepcogito/cogito-v2.1-671b"

ACCOUNTS = {
    "gurufe": {
        "account": "gurufe@gmail.com",
        "token_file": "token_drive_gurufe.json",
    },
    "pilates": {
        "account": "pilatesfisioiregua@gmail.com",
        "token_file": "token_drive_pilates.json",
    },
}

# Tipos de archivo que podemos leer
READABLE_MIMES = {
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "application/vnd.google-apps.document": "gdoc",
    "application/vnd.google-apps.spreadsheet": "gsheet",
    "text/plain": "txt",
    "text/markdown": "md",
    "text/csv": "csv",
}

# Google Docs export MIME types
EXPORT_MIMES = {
    "application/vnd.google-apps.document": "text/plain",
    "application/vnd.google-apps.spreadsheet": "text/csv",
    "application/vnd.google-apps.presentation": "text/plain",
}

MAX_TEXT_CHARS = 3000  # Max chars per document to send to LLM


# ─── AUTH ─────────────────────────────────────────────────────────────────────

def authenticate(account_key: str) -> Any:
    config = ACCOUNTS[account_key]
    token_path = SCRIPT_DIR / config["token_file"]
    creds = None

    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not CREDENTIALS_FILE.exists():
                print(f"ERROR: No se encuentra {CREDENTIALS_FILE}")
                sys.exit(1)
            print(f"\n  Autenticando {config['account']} (Drive)...")
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS_FILE), SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, "w") as f:
            f.write(creds.to_json())

    return build("drive", "v3", credentials=creds)


# ─── TEXT EXTRACTION ──────────────────────────────────────────────────────────

def extract_text_pdf(content: bytes) -> str:
    """Extract text from PDF bytes."""
    import pdfplumber
    text = ""
    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages[:10]:  # Max 10 pages
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
                if len(text) > MAX_TEXT_CHARS:
                    break
    except Exception:
        pass
    return text[:MAX_TEXT_CHARS]


def extract_text_docx(content: bytes) -> str:
    """Extract text from DOCX bytes."""
    from docx import Document
    try:
        doc = Document(io.BytesIO(content))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return text[:MAX_TEXT_CHARS]
    except Exception:
        return ""


def extract_text_xlsx(content: bytes) -> str:
    """Extract text from XLSX bytes."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        text = ""
        for sheet in wb.sheetnames[:3]:  # Max 3 sheets
            ws = wb[sheet]
            text += f"[Hoja: {sheet}]\n"
            for row in ws.iter_rows(max_row=50, values_only=True):
                vals = [str(v) for v in row if v is not None]
                if vals:
                    text += " | ".join(vals) + "\n"
                if len(text) > MAX_TEXT_CHARS:
                    break
        return text[:MAX_TEXT_CHARS]
    except Exception:
        return ""


def extract_text_plain(content: bytes) -> str:
    """Extract text from plain text files."""
    try:
        text = content.decode("utf-8", errors="replace")
        return text[:MAX_TEXT_CHARS]
    except Exception:
        return ""


EXTRACTORS = {
    "pdf": extract_text_pdf,
    "docx": extract_text_docx,
    "xlsx": extract_text_xlsx,
    "txt": extract_text_plain,
    "md": extract_text_plain,
    "csv": extract_text_plain,
}


# ─── DRIVE HELPERS ────────────────────────────────────────────────────────────

def list_all_files(service: Any) -> List[dict]:
    """List all readable files in Drive."""
    all_files = []
    page_token = None

    while True:
        kwargs = {
            "pageSize": 1000,
            "fields": "nextPageToken, files(id, name, mimeType, parents, size, modifiedTime, webViewLink)",
            "q": "trashed = false",
            "spaces": "drive",
        }
        if page_token:
            kwargs["pageToken"] = page_token

        results = service.files().list(**kwargs).execute()
        all_files.extend(results.get("files", []))
        page_token = results.get("nextPageToken")
        if not page_token:
            break

    return all_files


def download_file_content(service: Any, file_info: dict) -> Optional[bytes]:
    """Download file content. Returns bytes or None."""
    mime = file_info["mimeType"]
    file_id = file_info["id"]

    try:
        if mime in EXPORT_MIMES:
            # Google Docs/Sheets: export
            export_mime = EXPORT_MIMES[mime]
            response = service.files().export(fileId=file_id, mimeType=export_mime).execute()
            if isinstance(response, bytes):
                return response
            return response.encode("utf-8") if isinstance(response, str) else None
        else:
            # Regular file: download
            response = service.files().get_media(fileId=file_id).execute()
            return response
    except Exception as e:
        return None


def get_folder_path(service: Any, file_info: dict, folder_cache: dict) -> str:
    """Get the full folder path for a file."""
    parents = file_info.get("parents", [])
    if not parents:
        return "/"

    parent_id = parents[0]
    if parent_id in folder_cache:
        return folder_cache[parent_id]

    # Build path by traversing parents
    path_parts = []
    current_id = parent_id
    seen = set()

    while current_id and current_id not in seen:
        seen.add(current_id)
        if current_id in folder_cache:
            path_parts.insert(0, folder_cache[current_id])
            break
        try:
            folder = service.files().get(fileId=current_id, fields="name, parents").execute()
            path_parts.insert(0, folder.get("name", ""))
            parents = folder.get("parents", [])
            current_id = parents[0] if parents else None
        except Exception:
            break

    full_path = "/" + "/".join(path_parts) if path_parts else "/"
    folder_cache[parent_id] = full_path
    return full_path


# ─── LLM ──────────────────────────────────────────────────────────────────────

def call_llm(model: str, system: str, user_msg: str, max_tokens: int = 500) -> Optional[str]:
    """Call OpenRouter LLM via curl."""
    payload = json.dumps({
        "model": model,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user_msg},
        ],
        "max_tokens": max_tokens,
        "temperature": 0.3,
    })

    result = subprocess.run(
        [
            "curl", "-s", "https://openrouter.ai/api/v1/chat/completions",
            "-H", "Content-Type: application/json",
            "-H", f"Authorization: Bearer {OPENROUTER_KEY}",
            "-d", payload,
        ],
        capture_output=True, text=True, timeout=120,
    )

    try:
        data = json.loads(result.stdout)
        content = data["choices"][0]["message"]["content"]
        # Strip think tags
        content = re.sub(r"<think>.*?</think>", "", content, flags=re.DOTALL).strip()
        return content
    except Exception as e:
        return None


# ─── PHASE 1+2: CRAWL + ANALYZE ──────────────────────────────────────────────

ANALYSIS_SYSTEM = """Eres un archivista experto. Analiza el contenido de este documento y responde SOLO en JSON:

{
  "tema": "tema principal en 3-5 palabras",
  "tags": ["tag1", "tag2", "tag3"],
  "tipo": "factura|contrato|formación|personal|negocio|tech|legal|financiero|marketing|diseño|otro",
  "idioma": "es|en",
  "resumen": "una frase describiendo el contenido"
}

Si el texto está vacío o es ilegible, responde: {"tema": "ilegible", "tags": [], "tipo": "otro", "idioma": "?", "resumen": "no se pudo extraer texto"}"""


def analyze_documents(account_key: str) -> None:
    """Phase 1+2: Crawl Drive and analyze each document with LLM."""
    config = ACCOUNTS[account_key]
    account = config["account"]
    data_file = DATA_DIR / f"analysis_{account_key}.json"

    print(f"\n{'='*60}")
    print(f" Análisis semántico: {account}")
    print(f"{'='*60}")

    service = authenticate(account_key)

    # Load existing analysis (resume capability)
    existing = {}
    if data_file.exists():
        with open(data_file) as f:
            existing = json.load(f)
        print(f"  Análisis previo: {len(existing)} documentos")

    # List all files
    print(f"  Listando archivos...")
    all_files = list_all_files(service)

    # Filter readable files
    readable = []
    for f in all_files:
        mime = f["mimeType"]
        if mime in READABLE_MIMES or mime in EXPORT_MIMES:
            readable.append(f)

    folders = [f for f in all_files if f["mimeType"] == "application/vnd.google-apps.folder"]
    print(f"  Total archivos: {len(all_files)}")
    print(f"  Documentos legibles: {len(readable)}")
    print(f"  Carpetas: {len(folders)}")

    # Build folder cache
    folder_cache = {}
    for folder in folders:
        folder_cache[folder["id"]] = folder["name"]

    # Analyze each document
    new_count = 0
    skip_count = 0
    error_count = 0

    for i, file_info in enumerate(readable, 1):
        file_id = file_info["id"]
        name = file_info["name"]
        mime = file_info["mimeType"]

        # Skip if already analyzed
        if file_id in existing:
            skip_count += 1
            continue

        # Determine file type
        file_type = READABLE_MIMES.get(mime) or "gdoc" if "google-apps" in mime else None
        if not file_type:
            continue

        # Download content
        content = download_file_content(service, file_info)
        if not content:
            existing[file_id] = {
                "name": name,
                "mime": mime,
                "path": get_folder_path(service, file_info, folder_cache),
                "analysis": {"tema": "sin contenido", "tags": [], "tipo": "otro", "resumen": "no se pudo descargar"},
                "error": True,
            }
            error_count += 1
            continue

        # Extract text
        extractor = EXTRACTORS.get(file_type)
        if file_type in ("gdoc", "gsheet"):
            extractor = extract_text_plain

        text = extractor(content) if extractor else ""

        if not text.strip():
            existing[file_id] = {
                "name": name,
                "mime": mime,
                "path": get_folder_path(service, file_info, folder_cache),
                "analysis": {"tema": "sin texto", "tags": [], "tipo": "otro", "resumen": "documento sin texto extraíble"},
                "error": True,
            }
            error_count += 1
            continue

        # Analyze with LLM
        prompt = f"Archivo: {name}\nRuta actual: {get_folder_path(service, file_info, folder_cache)}\n\nContenido:\n{text}"

        print(f"  [{i}/{len(readable)}] Analizando: {name[:50]}...", end=" ", flush=True)

        llm_response = call_llm(ANALYSIS_MODEL, ANALYSIS_SYSTEM, prompt)

        if llm_response:
            try:
                # Try to parse JSON from response
                json_match = re.search(r'\{[^{}]*\}', llm_response, re.DOTALL)
                if json_match:
                    analysis = json.loads(json_match.group())
                else:
                    analysis = {"tema": "parse error", "tags": [], "tipo": "otro", "resumen": llm_response[:100]}
            except json.JSONDecodeError:
                analysis = {"tema": "parse error", "tags": [], "tipo": "otro", "resumen": llm_response[:100]}

            existing[file_id] = {
                "name": name,
                "mime": mime,
                "size": int(file_info.get("size", 0)),
                "modified": file_info.get("modifiedTime", ""),
                "path": get_folder_path(service, file_info, folder_cache),
                "parents": file_info.get("parents", []),
                "analysis": analysis,
            }
            new_count += 1
            print(f"→ {analysis.get('tipo', '?')} [{', '.join(analysis.get('tags', [])[:3])}]")
        else:
            error_count += 1
            print("→ ERROR LLM")

        # Save periodically
        if new_count % 10 == 0:
            with open(data_file, "w") as f:
                json.dump(existing, f, indent=2, ensure_ascii=False)

        # Rate limit
        time.sleep(0.5)

    # Final save
    with open(data_file, "w") as f:
        json.dump(existing, f, indent=2, ensure_ascii=False)

    print(f"\n── Resumen análisis ──")
    print(f"  Nuevos analizados: {new_count}")
    print(f"  Ya existentes:     {skip_count}")
    print(f"  Errores:           {error_count}")
    print(f"  Guardado en:       {data_file}")

    # Stats by type
    by_type = defaultdict(int)
    for doc in existing.values():
        tipo = doc.get("analysis", {}).get("tipo", "otro")
        by_type[tipo] += 1

    print(f"\n  Por tipo semántico:")
    for tipo, count in sorted(by_type.items(), key=lambda x: -x[1]):
        print(f"    {tipo}: {count}")


# ─── PHASE 3: PROPOSE REORGANIZATION ─────────────────────────────────────────

PROPOSAL_SYSTEM = """Eres un experto en organización de archivos digitales. Te doy una lista de documentos con su análisis semántico.

Tu tarea: proponer una estructura de carpetas ÓPTIMA basada en el CONTENIDO (no en el nombre del archivo).

Reglas:
1. Máximo 2 niveles de profundidad (carpeta/subcarpeta)
2. Cada carpeta debe tener un nombre descriptivo y claro
3. No crear carpetas con menos de 2 archivos
4. Agrupar por tema/función, no por formato
5. Incluir una carpeta "_archivo" para documentos obsoletos o duplicados

Responde en JSON:
{
  "carpetas": [
    {
      "nombre": "Nombre de la carpeta",
      "subcarpetas": ["sub1", "sub2"],
      "descripcion": "qué va aquí",
      "archivos": ["file_id1", "file_id2"]
    }
  ],
  "razonamiento": "explicación breve de por qué esta estructura"
}"""


def propose_reorganization(account_key: str) -> None:
    """Phase 3: Use LLM to propose optimal folder structure."""
    data_file = DATA_DIR / f"analysis_{account_key}.json"
    proposal_file = DATA_DIR / f"proposal_{account_key}.json"

    if not data_file.exists():
        print(f"ERROR: Primero ejecuta --analyze para {account_key}")
        return

    with open(data_file) as f:
        analysis = json.load(f)

    print(f"\n{'='*60}")
    print(f" Propuesta de reorganización: {ACCOUNTS[account_key]['account']}")
    print(f" Documentos analizados: {len(analysis)}")
    print(f"{'='*60}")

    # Build summary for LLM
    doc_summaries = []
    for file_id, doc in analysis.items():
        if doc.get("error"):
            continue
        a = doc.get("analysis", {})
        doc_summaries.append(
            f"- ID: {file_id}\n"
            f"  Archivo: {doc['name']}\n"
            f"  Ruta actual: {doc.get('path', '/')}\n"
            f"  Tipo: {a.get('tipo', '?')}\n"
            f"  Tema: {a.get('tema', '?')}\n"
            f"  Tags: {', '.join(a.get('tags', []))}\n"
            f"  Resumen: {a.get('resumen', '?')}"
        )

    if not doc_summaries:
        print("  No hay documentos para reorganizar.")
        return

    # Send to LLM in chunks if too many
    all_text = "\n\n".join(doc_summaries)
    print(f"  Enviando {len(doc_summaries)} documentos al LLM para propuesta...")
    print(f"  Modelo: {PROPOSAL_MODEL}")
    print(f"  Caracteres: {len(all_text)}")

    # If too large, truncate
    if len(all_text) > 80000:
        all_text = all_text[:80000] + "\n\n[... truncado por límite de contexto]"

    response = call_llm(
        PROPOSAL_MODEL,
        PROPOSAL_SYSTEM,
        f"Documentos a reorganizar:\n\n{all_text}",
        max_tokens=4000,
    )

    if not response:
        print("  ERROR: No se obtuvo respuesta del LLM")
        return

    # Try to parse JSON
    try:
        json_match = re.search(r'\{[\s\S]*\}', response)
        if json_match:
            proposal = json.loads(json_match.group())
        else:
            proposal = {"raw_response": response}
    except json.JSONDecodeError:
        proposal = {"raw_response": response}

    # Save proposal
    with open(proposal_file, "w") as f:
        json.dump(proposal, f, indent=2, ensure_ascii=False)

    # Display
    print(f"\n── Propuesta de reorganización ──\n")

    if "carpetas" in proposal:
        for folder in proposal["carpetas"]:
            name = folder.get("nombre", "?")
            desc = folder.get("descripcion", "")
            files = folder.get("archivos", [])
            subs = folder.get("subcarpetas", [])
            print(f"  📁 {name} ({len(files)} archivos)")
            if desc:
                print(f"     {desc}")
            if subs:
                for sub in subs:
                    print(f"     📁 {sub}/")

        if "razonamiento" in proposal:
            print(f"\n  Razonamiento: {proposal['razonamiento']}")
    else:
        print(response)

    print(f"\n  Propuesta guardada en: {proposal_file}")
    print(f"  Para ejecutar: python3 gdrive_semantic_agent.py --execute --account {account_key}")


# ─── PHASE 4: EXECUTE ────────────────────────────────────────────────────────

def execute_reorganization(account_key: str) -> None:
    """Phase 4: Execute the proposed reorganization."""
    proposal_file = DATA_DIR / f"proposal_{account_key}.json"
    data_file = DATA_DIR / f"analysis_{account_key}.json"

    if not proposal_file.exists():
        print(f"ERROR: Primero ejecuta --propose para {account_key}")
        return

    with open(proposal_file) as f:
        proposal = json.load(f)

    with open(data_file) as f:
        analysis = json.load(f)

    if "carpetas" not in proposal:
        print("ERROR: La propuesta no tiene estructura de carpetas válida")
        return

    config = ACCOUNTS[account_key]
    print(f"\n{'='*60}")
    print(f" Ejecutar reorganización: {config['account']}")
    print(f"{'='*60}")

    # Show summary
    total_moves = sum(len(f.get("archivos", [])) for f in proposal["carpetas"])
    print(f"\n  Carpetas a crear: {len(proposal['carpetas'])}")
    print(f"  Archivos a mover: {total_moves}")

    # Confirm
    print(f"\n  ⚠️  Esto moverá archivos en Google Drive.")
    confirm = input("  ¿Continuar? (s/n): ").strip().lower()
    if confirm != "s":
        print("  Cancelado.")
        return

    service = authenticate(account_key)

    # Create folders and move files
    moved = 0
    errors = 0

    for folder_spec in proposal["carpetas"]:
        folder_name = folder_spec.get("nombre", "Sin nombre")
        file_ids = folder_spec.get("archivos", [])
        subcarpetas = folder_spec.get("subcarpetas", [])

        if not file_ids:
            continue

        # Create folder
        folder_id = get_or_create_folder(service, folder_name)
        if not folder_id:
            print(f"  ERROR creando carpeta: {folder_name}")
            errors += len(file_ids)
            continue

        # Create subcarpetas
        sub_ids = {}
        for sub in subcarpetas:
            sub_id = get_or_create_folder(service, sub, parent_id=folder_id)
            if sub_id:
                sub_ids[sub] = sub_id

        # Move files
        for file_id in file_ids:
            doc = analysis.get(file_id, {})
            name = doc.get("name", file_id)

            try:
                # Get current parents
                file_meta = service.files().get(fileId=file_id, fields="parents").execute()
                current_parents = file_meta.get("parents", [])
                previous_parents = ",".join(current_parents)

                service.files().update(
                    fileId=file_id,
                    addParents=folder_id,
                    removeParents=previous_parents,
                    fields="id, parents",
                ).execute()

                moved += 1
                print(f"  ✓ {name[:50]} → {folder_name}/")
            except Exception as e:
                errors += 1
                print(f"  ✗ {name[:50]}: {e}")

            time.sleep(0.2)  # Rate limit

    print(f"\n── Resumen ejecución ──")
    print(f"  Archivos movidos: {moved}")
    print(f"  Errores:          {errors}")


def get_or_create_folder(service: Any, name: str, parent_id: str = None) -> Optional[str]:
    """Get or create a folder in Drive."""
    query = f"name = '{name}' and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    if parent_id:
        query += f" and '{parent_id}' in parents"

    results = service.files().list(q=query, fields="files(id)", spaces="drive").execute()
    files = results.get("files", [])

    if files:
        return files[0]["id"]

    metadata = {"name": name, "mimeType": "application/vnd.google-apps.folder"}
    if parent_id:
        metadata["parents"] = [parent_id]

    folder = service.files().create(body=metadata, fields="id").execute()
    print(f"  + Carpeta creada: {name}")
    return folder["id"]


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Agente semántico Google Drive")
    parser.add_argument("--analyze", action="store_true", help="Fase 1+2: crawl y analizar documentos")
    parser.add_argument("--propose", action="store_true", help="Fase 3: proponer reorganización")
    parser.add_argument("--execute", action="store_true", help="Fase 4: ejecutar reorganización")
    parser.add_argument("--account", choices=["gurufe", "pilates", "all"], default="all")
    args = parser.parse_args()

    if not any([args.analyze, args.propose, args.execute]):
        print("Especifica una fase: --analyze, --propose, o --execute")
        print("Ejemplo: python3 gdrive_semantic_agent.py --analyze --account gurufe")
        return

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    keys = list(ACCOUNTS.keys()) if args.account == "all" else [args.account]

    for key in keys:
        if args.analyze:
            analyze_documents(key)
        if args.propose:
            propose_reorganization(key)
        if args.execute:
            execute_reorganization(key)


if __name__ == "__main__":
    main()
